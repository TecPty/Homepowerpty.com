import openpyxl

path = r"c:\Users\HP 15\Homepowerpty.com\EXCEL DE PRODUCTOS PARA MKT.xlsx"
wb = openpyxl.load_workbook(path)
print("Sheets:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n--- Sheet: {sheet_name} ---")
    # Print first 5 rows to understand structure
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 8:
            break
        print(row)
