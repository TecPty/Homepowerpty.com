import openpyxl

path = r"c:\Users\HP 15\Homepowerpty.com\EXCEL DE PRODUCTOS PARA MKT.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== SHEET: {sheet_name} (max_row={ws.max_row}, max_col={ws.max_column}) ===")
    
    # Scan only non-empty cells
    found = []
    for row in ws.iter_rows(values_only=True):
        non_empty = [str(c).strip() for c in row if c is not None and str(c).strip()]
        if non_empty:
            found.append(non_empty)
    
    print(f"Total non-empty rows: {len(found)}")
    for r in found[:30]:
        print(" | ".join(r[:10]))  # first 10 columns max
